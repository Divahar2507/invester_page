"""
InvestorConnect Project Plan Generator (2 Months / 4 Members)
Generates a comprehensive 60-day project plan Excel report for a 4-member team
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

def create_project_plan():
    """Generate comprehensive project plan Excel file"""
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    create_overview_sheet(wb)
    create_daily_tasks_sheet(wb)
    create_member_sheet(wb, "Member 1 - Startup Lead", "Member 1: Startup Portal Lead (Full Stack)", get_member1_features())
    create_member_sheet(wb, "Member 2 - Investor FE", "Member 2: Investor Portal Frontend Lead", get_member2_features())
    create_member_sheet(wb, "Member 3 - Investor BE", "Member 3: Investor Portal Backend Lead", get_member3_features())
    create_member_sheet(wb, "Member 4 - Admin & Infra", "Member 4: Admin & Shared Services Lead", get_member4_features())
    create_progress_tracker_sheet(wb)
    
    # Save file
    filename = f"InvestorConnect_2Month_Plan_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Project plan generated: {filename}")
    return filename


def create_overview_sheet(wb):
    """Create project overview sheet"""
    ws = wb.create_sheet("Project Overview", 0)
    
    # Headers
    ws['A1'] = "InvestorConnect - 2 Month Development Plan"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    # Project Info
    info = [
        ["Project Name:", "InvestorConnect (InfiniteTechAI Platform)"],
        ["Duration:", "60 Days (2 Months)"],
        ["Team Size:", "4 Members"],
        ["Start Date:", datetime.now().strftime("%Y-%m-%d")],
        ["End Date:", (datetime.now() + timedelta(days=60)).strftime("%Y-%m-%d")],
        ["", ""],
        ["Team Structure:", ""],
        ["Member 1:", "Startup Portal Lead (Full Stack)"],
        ["Member 2:", "Investor Portal Frontend Lead"],
        ["Member 3:", "Investor Portal Backend Lead"],
        ["Member 4:", "Admin & Shared Services (Auth, DevOps, QA)"],
    ]
    
    row = 3
    for item in info:
        ws[f'A{row}'] = item[0]
        ws[f'A{row}'].font = Font(bold=True)
        if len(item) > 1:
            ws[f'B{row}'] = item[1]
        row += 1
    
    # Timeline Breakdown
    row += 2
    ws[f'A{row}'] = "Timeline Phases"
    ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    
    phases = [
        ["Month 1: Weeks 1-2", "Foundations & Core Auth", "Environment, DB, Auth System, Basic UI Layouts"],
        ["Month 1: Weeks 3-4", "Core Profiles & Pitches", "Startup/Investor Profiles, Pitch Feed, Basic Flow"],
        ["Month 2: Weeks 5-6", "Connectivity & Messaging", "Real-time Chat, Connections, Notifications, Watchlist"],
        ["Month 2: Weeks 7-8", "Advanced Features & Launch", "Admin Dashboard, Analytics, Security Audit, Deployment"],
    ]
    
    row += 1
    for phase in phases:
        ws[f'A{row}'] = phase[0]
        ws[f'B{row}'] = phase[1]
        ws[f'C{row}'] = phase[2]
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50


def create_daily_tasks_sheet(wb):
    """Create daily tasks breakdown sheet"""
    ws = wb.create_sheet("Daily Tasks (60 Days)")
    
    # Headers
    headers = ["Day", "Date", "Member 1 (Startup)", "Member 2 (Investor FE)", "Member 3 (Investor BE)", "Member 4 (Admin/Infra)", "Milestone"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Daily tasks generator
    tasks = get_60_day_tasks()
    
    start_date = datetime.now()
    for day, task_set in enumerate(tasks, 1):
        current_date = start_date + timedelta(days=day-1)
        row = day + 1
        
        ws.cell(row, 1, f"Day {day}")
        ws.cell(row, 2, current_date.strftime("%Y-%m-%d"))
        ws.cell(row, 3, task_set.get("member1", "Continued Dev"))
        ws.cell(row, 4, task_set.get("member2", "Continued Dev"))
        ws.cell(row, 5, task_set.get("member3", "Continued Dev"))
        ws.cell(row, 6, task_set.get("member4", "Continued Dev"))
        ws.cell(row, 7, task_set.get("milestone", ""))
        
        # Styling
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Weekend highlighting
            if current_date.weekday() >= 5:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20


def get_60_day_tasks():
    """Return 60 days of tasks"""
    tasks = []
    
    # --- MONTH 1 ---
    
    # Week 1: Setup & Auth
    tasks.append({"member1": "Frontend Repo Init, React+Vite Setup", "member2": "Investor Repo Init, UI Lib Installation", "member3": "Backend Repo Init, FastAPI Setup", "member4": "Docker Compose Setup, DB Config", "milestone": "Project Kickoff 🚀"})
    tasks.append({"member1": "Landing Page Initial Layout", "member2": "Investor Auth Pages (Login/Register)", "member3": "User Models & Database Schema", "member4": "Auth Service (JWT, OAuth base)", "milestone": ""})
    tasks.append({"member1": "Startup Auth Pages (Login/Register)", "member2": "Setup Tailwind & Design System", "member3": "Auth API Endpoints (Login/Reg)", "member4": "Email Service Integration (SendGrid/SMTP)", "milestone": ""})
    tasks.append({"member1": "Connect Auth to Backend", "member2": "Connect Auth to Backend", "member3": "Middleware & Error Handling", "member4": "CI/CD Pipeline Basic Setup", "milestone": ""})
    tasks.append({"member1": "Onboarding Wizard Flow (UI)", "member2": "Onboarding Wizard Flow (UI)", "member3": "Profile Creation API", "member4": "S3/Storage Setup for Files", "milestone": "Auth System Ready"})
    tasks.append({"member1": "Code Review & Refactor", "member2": "Code Review & Refactor", "member3": "Code Review & DB Optimization", "member4": "Infrastructure Monitoring Setup", "milestone": "Weekend Catchup"})
    tasks.append({"member1": "Rest/Planning", "member2": "Rest/Planning", "member3": "Rest/Planning", "member4": "Rest/Planning", "milestone": "Sunday"})

    # Week 2: Profiles & Dashboards
    tasks.append({"member1": "Startup Profile Page UI", "member2": "Investor Profile Page UI", "member3": "Profile CRUD Endpoints", "member4": "Admin User Management Dashboard", "milestone": ""})
    tasks.append({"member1": "Edit Profile Functionality", "member2": "Edit Profile Functionality", "member3": "File Upload API (Logos/Avatars)", "member4": "Image Optimization Service", "milestone": ""})
    tasks.append({"member1": "Startup Dashboard Shell", "member2": "Investor Dashboard Shell", "member3": "Dashboard Stats API (Mock Data)", "member4": "Logging & Auditing System", "milestone": ""})
    tasks.append({"member1": "Integration of Profile API", "member2": "Integration of Profile API", "member3": "Dashboard Logic Implementation", "member4": "Verification Flow (Email/Phone)", "milestone": ""})
    tasks.append({"member1": "Pitch Creation Form UI (Step 1-2)", "member2": "Browse Startups Grid UI", "member3": "Pitch Models & Creation API", "member4": "Admin Approval Workflow Backend", "milestone": "Profiles Complete"})
    tasks.append({"member1": "Testing Profile Flows", "member2": "Testing Profile Flows", "member3": "Unit Tests for Profiles", "member4": "Security Headers & CORS Config", "milestone": ""})
    tasks.append({"member1": "Rest", "member2": "Rest", "member3": "Rest", "member4": "Rest", "milestone": ""})

    # Week 3: Pitches & Discovery
    tasks.append({"member1": "Pitch Form (Financials/Docs)", "member2": "Advanced Filters (Industry/Stage)", "member3": "Pitch Upload & Document API", "member4": "Document Security & Encryption", "milestone": ""})
    tasks.append({"member1": "Pitch Preview & Submit", "member2": "Detailed Startup View Modal", "member3": "Pitch Query/Filter API", "member4": "Admin Pitch Review Interface", "milestone": ""})
    tasks.append({"member1": "My Pitches Management UI", "member2": "Investment Logging Form UI", "member3": "Investment Tracking API", "member4": "Data Analytics Base Setup", "milestone": ""})
    tasks.append({"member1": "Pitch Analytics View", "member2": "Portfolio Dashboard UI", "member3": "Portfolio Logic & Calculations", "member4": "Report Generation Service (PDF)", "milestone": ""})
    tasks.append({"member1": "Integration of Pitch API", "member2": "Integration of Feed API", "member3": "Feed Algorithm (Basic)", "member4": "Matching Algorithm (Basic)", "milestone": "Pitch Flow Ready"})
    tasks.append({"member1": "Fix UI Bugs (Forms)", "member2": "Fix UI Bugs (Grids)", "member3": "Optimize Feed Query", "member4": "Load Testing Feed", "milestone": ""})
    tasks.append({"member1": "Rest", "member2": "Rest", "member3": "Rest", "member4": "Rest", "milestone": ""})

    # Week 4: Messaging Foundation
    tasks.append({"member1": "Chat List UI", "member2": "Chat List UI", "member3": "Message Models & Basic API", "member4": "WebSocket Server Setup", "milestone": ""})
    tasks.append({"member1": "Chat Room/Conversation UI", "member2": "Chat Room/Conversation UI", "member3": "Send/Receive API Endpoints", "member4": "Real-time Event Broadcasting", "milestone": ""})
    tasks.append({"member1": "Connect WS to Frontend", "member2": "Connect WS to Frontend", "member3": "WS Authentication Handler", "member4": "Redis for Pub/Sub (Scalability)", "milestone": ""})
    tasks.append({"member1": "Message States (IsTyping/Read)", "member2": "Message States (IsTyping/Read)", "member3": "Read Receipts Backend", "member4": "Message Persistence & History", "milestone": ""})
    tasks.append({"member1": "File Sharing in Chat UI", "member2": "File Sharing in Chat UI", "member3": "Chat File Sharing API", "member4": "Malware Scanning Service", "milestone": "Messaging MVP"})
    tasks.append({"member1": "Integration Testing Chat", "member2": "Integration Testing Chat", "member3": "WS Stress Testing", "member4": "Server Scaling Config", "milestone": "Month 1 Review"})
    tasks.append({"member1": "Rest", "member2": "Rest", "member3": "Rest", "member4": "Rest", "milestone": ""})

    # --- MONTH 2 ---

    # Week 5: Connections & Social
    tasks.append({"member1": "Connection Request UI", "member2": "Find Investors Page", "member3": "Connections API & Network Graph", "member4": "Notification Service (Brain)", "milestone": ""})
    tasks.append({"member1": "Network Management Page", "member2": "Connection Request Handling", "member3": "Connection Logic & Constraints", "member4": "Push Notification Integration", "milestone": ""})
    tasks.append({"member1": "Notifications Dropdown UI", "member2": "Notifications Dropdown UI", "member3": "Notification Trigger System", "member4": "Email Digests System", "milestone": ""})
    tasks.append({"member1": "In-App Notification Stream", "member2": "In-App Notification Stream", "member3": "Real-time Notification WS", "member4": "Activity Logs System", "milestone": ""})
    tasks.append({"member1": "Settings: Notifications", "member2": "Settings: Notifications", "member3": "Preferences API", "member4": "GDPR/Compliance Tools", "milestone": "Social Features Done"})
    tasks.append({"member1": "Catchup / Polish", "member2": "Catchup / Polish", "member3": "Catchup / DB Migrations", "member4": "Backup System Test", "milestone": ""})
    tasks.append({"member1": "Rest", "member2": "Rest", "member3": "Rest", "member4": "Rest", "milestone": ""})

    # Week 6: Advanced Tools
    tasks.append({"member1": "Data Room Manager UI", "member2": "Data Room Viewer UI", "member3": "Data Room Permissions API", "member4": "Access Logs & Audit Trails", "milestone": ""})
    tasks.append({"member1": "Upload Financials Logic", "member2": "Download Tracking Logic", "member3": "Secure File Delivery API", "member4": "Watermarking Service", "milestone": ""})
    tasks.append({"member1": "Meeting Scheduler Integration", "member2": "Calendar View Integration", "member3": "Meeting API & Calendar Sync", "member4": "Google/Outlook Calendar Integration", "milestone": ""})
    tasks.append({"member1": "Task/To-Do List Widget", "member2": "Watchlist & Notes Widget", "member3": "Tasks & Notes API", "member4": "Search Engine (Elastic/PGSearch)", "milestone": ""})
    tasks.append({"member1": "Global Search Bar UI", "member2": "Global Search Bar UI", "member3": "Global Search API", "member4": "Search Indexing Service", "milestone": "Tools Complete"})
    tasks.append({"member1": "Review Tools UX", "member2": "Review Tools UX", "member3": "API Performance Tuning", "member4": "DB Index optimization", "milestone": ""})
    tasks.append({"member1": "Rest", "member2": "Rest", "member3": "Rest", "member4": "Rest", "milestone": ""})

    # Week 7: Admin & Analytics
    tasks.append({"member1": "Startup Analytics Dashboard", "member2": "Investor AI Recommendations", "member3": "Analytics Aggregation API", "member4": "Super Admin Dashboard Setup", "milestone": ""})
    tasks.append({"member1": "Export Reports Function", "member2": "Export Reports Function", "member3": "Detailed Report Generation API", "member4": "Admin User Moderation Tools", "milestone": ""})
    tasks.append({"member1": "Feedback/Support Form", "member2": "Feedback/Support Form", "member3": "Support Ticket API", "member4": "Admin Support Queue", "milestone": ""})
    tasks.append({"member1": "Polish: Loading States", "member2": "Polish: Empty States", "member3": "Backend Caching Layer", "member4": "System Health Dashboard", "milestone": ""})
    tasks.append({"member1": "Polish: Error Boundaries", "member2": "Polish: Animations", "member3": "API Rate Limiting Finals", "member4": "Security Pen-Testing (Self)", "milestone": "Feature Freeze"})
    tasks.append({"member1": "Bug Fixing", "member2": "Bug Fixing", "member3": "Bug Fixing", "member4": "Bug Fixing", "milestone": ""})
    tasks.append({"member1": "Rest", "member2": "Rest", "member3": "Rest", "member4": "Rest", "milestone": ""})

    # Week 8: Launch Prep
    tasks.append({"member1": "End-to-End Testing (User Flow)", "member2": "End-to-End Testing (User Flow)", "member3": "Final API Tests", "member4": "Production Environment Setup", "milestone": "Start QA"})
    tasks.append({"member1": "Fix Critical Bugs", "member2": "Fix Critical Bugs", "member3": "Fix Critical Bugs", "member4": "SSL & Domain Config", "milestone": ""})
    tasks.append({"member1": "Final UI Polish", "member2": "Final UI Polish", "member3": "Database Final Migration", "member4": "CDN & Asset Optimization", "milestone": ""})
    tasks.append({"member1": "Documentation (User Manual)", "member2": "Documentation (Videos)", "member3": "API Docs Finalizing", "member4": "Deployment Runbook", "milestone": ""})
    tasks.append({"member1": "Pre-Launch Check", "member2": "Pre-Launch Check", "member3": "Pre-Launch Check", "member4": "Go-Live execution", "milestone": "GO LIVE 🚀"})
    tasks.append({"member1": "Post-Launch Monitoring", "member2": "Post-Launch Monitoring", "member3": "Live Log Monitoring", "member4": "Incident Response Watch", "milestone": ""})
    tasks.append({"member1": "Victory Lap", "member2": "Victory Lap", "member3": "Victory Lap", "member4": "Victory Lap", "milestone": "PROJECT COMPLETE"})
    
    return tasks


def create_member_sheet(wb, sheet_name, title, features):
    """Create individual member task sheet"""
    ws = wb.create_sheet(sheet_name)
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    ws.merge_cells('A1:E1')
    
    headers = ["Component", "Detailed Tasks", "Priority", "Est. Hours", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    
    row = 4
    for feature in features:
        ws.cell(row, 1, feature["name"])
        ws.cell(row, 2, feature["tasks"])
        ws.cell(row, 3, feature["priority"])
        ws.cell(row, 4, 8) # Placeholder hours
        ws.cell(row, 5, "Pending")
        
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        
        # Priority color coding
        priority_colors = {"High": "FF6B6B", "Medium": "FFA500", "Low": "4ECDC4"}
        ws.cell(row, 3).fill = PatternFill(start_color=priority_colors.get(feature["priority"], "FFFFFF"), 
                                           end_color=priority_colors.get(feature["priority"], "FFFFFF"), 
                                           fill_type="solid")
        
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15


def get_member1_features():
    """Startup Portal Lead"""
    return [
        {"name": "Auth & Onboarding", "tasks": "Signup, Login, Wizard Flow, Email Verification", "priority": "High"},
        {"name": "Startup Profile", "tasks": "Profile Form, Team Members, Logo Upload, Social Links", "priority": "High"},
        {"name": "Pitch Management", "tasks": "Create Pitch Flow, Edit Pitch, Document Uploads", "priority": "High"},
        {"name": "Dashboard", "tasks": "Analytics Widgets, Recent Visitors, Notifications Feed", "priority": "High"},
        {"name": "Investor Search", "tasks": "Search Interface, Filters, Investor Profiles View", "priority": "Medium"},
        {"name": "Messaging UI", "tasks": "Chat Interface, File Sharing, Contacts List", "priority": "High"},
        {"name": "Settings", "tasks": "Account Settings, Privacy, Notifications Preferences", "priority": "Low"},
    ]

def get_member2_features():
    """Investor Portal Frontend Lead"""
    return [
        {"name": "Auth & Onboarding", "tasks": "Investor Signup, KYC Flow, Thesis Definition", "priority": "High"},
        {"name": "Discovery", "tasks": "Startup Feed, Specialized Filters, Search, Recommendation Cards", "priority": "High"},
        {"name": "Due Diligence", "tasks": "Startup Details Modal, Data Room Viewer, Q&A Section", "priority": "High"},
        {"name": "Portfolio", "tasks": "My Investments, Performance Charts, Add Investment Flow", "priority": "High"},
        {"name": "Pipeline", "tasks": "Kanban Board for Deal Flow, Notes, Status Updates", "priority": "Medium"},
        {"name": "Social", "tasks": "Networking Interface, Messaging, Calendar/Meetings", "priority": "Medium"},
        {"name": "Reporting", "tasks": "Export to PDF/Excel, Portfolio Summary", "priority": "Low"},
    ]

def get_member3_features():
    """Investor Portal Backend Lead"""
    return [
        {"name": "Core API", "tasks": "User CRUD, Authentication, Profile Management", "priority": "High"},
        {"name": "Pitch Engine", "tasks": "Pitch CRUD, Feed Algorithm, Search Optimization", "priority": "High"},
        {"name": "Matching Logic", "tasks": "Score Calculation, Recommendation Engine", "priority": "High"},
        {"name": "Investment API", "tasks": "Portfolio Tracking, Transaction Records, Analytics", "priority": "High"},
        {"name": "Communication", "tasks": "Messaging Endpoints, Notification Triggers", "priority": "High"},
        {"name": "Data Room", "tasks": "Secure File Upload/Download, Access Control Lists", "priority": "Medium"},
        {"name": "Integrations", "tasks": "Calendar Sync, External Data Sources", "priority": "Low"},
    ]

def get_member4_features():
    """Admin & Shared Services Lead"""
    return [
        {"name": "Infrastructure", "tasks": "Docker Setup, CI/CD, Database Management, Backups", "priority": "High"},
        {"name": "Security", "tasks": "Auth Service (OAuth), Rate Limiting, CORS, Encryption", "priority": "High"},
        {"name": "Admin Dashboard", "tasks": "User Management, Content Moderation, Analytics", "priority": "High"},
        {"name": "Notification Svc", "tasks": "Push/Email Service, Template Management, Queues", "priority": "Medium"},
        {"name": "Real-time Svc", "tasks": "WebSocket Server, Redis Integration, Event Bus", "priority": "High"},
        {"name": "DevOps", "tasks": "Monitoring (Prometheus/Grafana), Logging (ELK)", "priority": "Medium"},
        {"name": "Compliance", "tasks": "GDPR Tools, Data Export, Audit Logging", "priority": "Low"},
    ]

def create_progress_tracker_sheet(wb):
    """Create progress tracking sheet"""
    ws = wb.create_sheet("Progress Tracker")
    
    ws['A1'] = "Weekly Progress Tracker"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    ws.merge_cells('A1:E1')
    
    headers = ["Week", "Planned Tasks", "Completed Tasks", "Completion %", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    
    for i in range(1, 9):
        row = i + 3
        ws.cell(row, 1, f"Week {i}")
        ws.cell(row, 2, 28) # Avg tasks per week (4 people * 7 days)
        ws.cell(row, 3, 0)
        ws.cell(row, 4, "=C{}/B{}*100".format(row, row))
        ws.cell(row, 5, "")
        ws.cell(row, 4).number_format = '0.00"%"'
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40


if __name__ == "__main__":
    print("🚀 Generating InvestorConnect 2-Month Project Plan...")
    print("=" * 60)
    filename = create_project_plan()
    print("=" * 60)
    print(f"📊 Excel file created successfully!")
    print(f"📁 Location: {os.path.abspath(filename)}")
    print("\n💡 Open the file in Excel to view: Project Overview, Daily Tasks, and Individual Assignments.")
