"""
InvestorConnect Project Plan Generator
Generates a comprehensive 1-month project plan Excel report for a 3-member team
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

def create_project_plan():
    """Generate comprehensive project plan Excel file"""
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    create_overview_sheet(wb)
    create_daily_tasks_sheet(wb)
    create_member1_tasks_sheet(wb)  # Startup Portal
    create_member2_tasks_sheet(wb)  # Investor Frontend
    create_member3_tasks_sheet(wb)  # Investor Backend
    create_progress_tracker_sheet(wb)
    
    # Save file
    filename = f"InvestorConnect_ProjectPlan_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Project plan generated: {filename}")
    return filename


def create_overview_sheet(wb):
    """Create project overview sheet"""
    ws = wb.create_sheet("Project Overview", 0)
    
    # Headers
    ws['A1'] = "InvestorConnect - 1 Month Development Plan"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    # Project Info
    info = [
        ["Project Name:", "InvestorConnect (InfiniteTechAI Platform)"],
        ["Duration:", "30 Days"],
        ["Team Size:", "3 Members"],
        ["Start Date:", datetime.now().strftime("%Y-%m-%d")],
        ["End Date:", (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")],
        ["", ""],
        ["Team Structure:", ""],
        ["Member 1:", "Startup Portal (Frontend + Backend)"],
        ["Member 2:", "Investor Portal (Frontend)"],
        ["Member 3:", "Investor Portal (Backend + Infrastructure)"],
    ]
    
    row = 3
    for item in info:
        ws[f'A{row}'] = item[0]
        ws[f'A{row}'].font = Font(bold=True)
        if len(item) > 1:
            ws[f'B{row}'] = item[1]
        row += 1
    
    # Weekly Breakdown
    row += 2
    ws[f'A{row}'] = "Weekly Breakdown"
    ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    
    weeks = [
        ["Week 1 (Days 1-7)", "Foundation & Core Features", "Setup, Auth, Basic UI"],
        ["Week 2 (Days 8-14)", "Advanced Features", "Messaging, Portfolio, Data Rooms"],
        ["Week 3 (Days 15-21)", "Integration & Testing", "API Integration, Bug Fixes, Testing"],
        ["Week 4 (Days 22-30)", "Deployment & Polish", "Production Deploy, Documentation, Final QA"],
    ]
    
    row += 1
    for week in weeks:
        ws[f'A{row}'] = week[0]
        ws[f'B{row}'] = week[1]
        ws[f'C{row}'] = week[2]
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40


def create_daily_tasks_sheet(wb):
    """Create daily tasks breakdown sheet"""
    ws = wb.create_sheet("Daily Tasks")
    
    # Headers
    headers = ["Day", "Date", "Member 1 (Startup)", "Member 2 (Investor FE)", "Member 3 (Investor BE)", "Milestone"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Daily tasks
    tasks = get_daily_tasks()
    
    start_date = datetime.now()
    for day, task_set in enumerate(tasks, 1):
        current_date = start_date + timedelta(days=day-1)
        row = day + 1
        
        ws.cell(row, 1, f"Day {day}")
        ws.cell(row, 2, current_date.strftime("%Y-%m-%d"))
        ws.cell(row, 3, task_set["member1"])
        ws.cell(row, 4, task_set["member2"])
        ws.cell(row, 5, task_set["member3"])
        ws.cell(row, 6, task_set.get("milestone", ""))
        
        # Styling
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Weekend highlighting
            if current_date.weekday() >= 5:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 25


def get_daily_tasks():
    """Return 30 days of tasks"""
    return [
        # Week 1: Foundation
        {"member1": "Setup Startup Portal repo, Docker config, Basic routing", 
         "member2": "Setup Investor Portal repo, Install dependencies, Create layout",
         "member3": "Database schema design, Setup PostgreSQL, Create base models",
         "milestone": "Environment Setup"},
        
        {"member1": "Create Signup page UI, Form validation, Error handling", 
         "member2": "Create Login page UI, Design system setup, Color palette",
         "member3": "Auth endpoints (register, login), JWT implementation, Password hashing",
         "milestone": ""},
        
        {"member1": "Create Login page, Integrate auth API, Token storage", 
         "member2": "Dashboard layout, Sidebar navigation, Header component",
         "member3": "User profile endpoints, Role-based access control, Session management",
         "milestone": "Auth Complete"},
        
        {"member1": "Dashboard UI, Stats cards, Recent activity feed", 
         "member2": "Browse Pitches page layout, Filter UI, Search bar",
         "member3": "Pitch feed API, Filtering logic, Search implementation",
         "milestone": ""},
        
        {"member1": "Company Profile page, Edit mode, Form handling", 
         "member2": "Pitch detail modal, Comments section UI, Rating stars",
         "member3": "Startup profile CRUD, Pitch CRUD endpoints, File upload setup",
         "milestone": "Core Pages Done"},
        
        {"member1": "Pitch creation form, File upload UI, Preview mode", 
         "member2": "Portfolio page layout, Investment cards, Stats dashboard",
         "member3": "Investment tracking API, Portfolio analytics, Data aggregation",
         "milestone": ""},
        
        {"member1": "Review & fix Week 1 bugs, Code cleanup, Documentation", 
         "member2": "Review & polish UI, Responsive fixes, Accessibility",
         "member3": "Database optimization, API performance tuning, Error handling",
         "milestone": "Week 1 Review"},
        
        # Week 2: Advanced Features
        {"member1": "Messages page UI, Conversation list, Chat interface", 
         "member2": "Messages page UI (investor side), User search, Chat UI",
         "member3": "Messaging API, WebSocket setup, Real-time notifications",
         "milestone": "Messaging Start"},
        
        {"member1": "File sharing in messages, Image preview, PDF support", 
         "member2": "Watchlist page, Add/remove functionality, Watchlist cards",
         "member3": "Watchlist API, Connection requests API, Status management",
         "milestone": ""},
        
        {"member1": "Notifications UI, Notification bell, Mark as read", 
         "member2": "In Review page, Pipeline view, Drag-drop (optional)",
         "member3": "Notifications backend, Push notifications, Email triggers",
         "milestone": "Notifications Done"},
        
        {"member1": "Investor discovery page, Filter by criteria, Connect button", 
         "member2": "Data Room component, Document viewer, Download tracking",
         "member3": "Data room API, Document storage, Access control",
         "milestone": ""},
        
        {"member1": "Settings page, Profile photo upload, Password change", 
         "member2": "Meeting scheduler UI, Calendar integration, Time slots",
         "member3": "Meeting API, Google Calendar integration, Email reminders",
         "milestone": "Settings Complete"},
        
        {"member1": "Analytics dashboard, Charts, Export functionality", 
         "member2": "Log Investment page, Form validation, Document upload",
         "member3": "Investment logging API, Portfolio stats, Growth tracking",
         "milestone": ""},
        
        {"member1": "Review & fix Week 2 bugs, Integration testing, Polish", 
         "member2": "Review & UI refinement, Cross-browser testing, Mobile fixes",
         "member3": "API testing, Load testing, Security audit",
         "milestone": "Week 2 Review"},
        
        # Week 3: Integration & Testing
        {"member1": "End-to-end testing (Signup → Pitch creation)", 
         "member2": "End-to-end testing (Browse → Investment logging)",
         "member3": "Integration testing, API documentation, Postman collection",
         "milestone": "Testing Phase"},
        
        {"member1": "Fix critical bugs, Edge case handling, Error messages", 
         "member2": "Fix UI bugs, Loading states, Empty states",
         "member3": "Fix backend bugs, Database migrations, Data validation",
         "milestone": ""},
        
        {"member1": "Performance optimization, Code splitting, Lazy loading", 
         "member2": "Performance optimization, Image optimization, Caching",
         "member3": "Database indexing, Query optimization, Caching layer",
         "milestone": "Performance Tuning"},
        
        {"member1": "Security review, XSS prevention, CSRF tokens", 
         "member2": "Security review, Input sanitization, Auth checks",
         "member3": "Security hardening, SQL injection prevention, Rate limiting",
         "milestone": ""},
        
        {"member1": "User acceptance testing, Feedback collection, Bug fixes", 
         "member2": "User acceptance testing, UI/UX improvements, Polish",
         "member3": "Monitoring setup, Logging, Error tracking (Sentry)",
         "milestone": "UAT Complete"},
        
        {"member1": "Documentation, User guide, Feature walkthrough", 
         "member2": "Documentation, Component library, Style guide",
         "member3": "API documentation, Deployment guide, Architecture docs",
         "milestone": ""},
        
        {"member1": "Review & final polish, Regression testing, Sign-off", 
         "member2": "Review & final polish, Accessibility audit, Final QA",
         "member3": "Review & final checks, Backup strategy, Rollback plan",
         "milestone": "Week 3 Review"},
        
        # Week 4: Deployment & Polish
        {"member1": "Production build, Environment variables, Config", 
         "member2": "Production build, Asset optimization, CDN setup",
         "member3": "Production database setup, Migrations, Seed data",
         "milestone": "Deployment Prep"},
        
        {"member1": "Deploy to staging, Smoke testing, Bug fixes", 
         "member2": "Deploy to staging, Cross-browser testing, Mobile testing",
         "member3": "Deploy backend to staging, Health checks, Monitoring",
         "milestone": ""},
        
        {"member1": "Production deployment, DNS setup, SSL certificates", 
         "member2": "Production deployment, CDN configuration, Analytics",
         "member3": "Production deployment, Database backups, Scaling config",
         "milestone": "Production Live"},
        
        {"member1": "Post-deployment monitoring, Bug fixes, Hot fixes", 
         "member2": "Post-deployment monitoring, User feedback, Quick fixes",
         "member3": "Post-deployment monitoring, Performance monitoring, Alerts",
         "milestone": ""},
        
        {"member1": "Feature enhancements, User feedback implementation", 
         "member2": "UI polish, Animation improvements, Micro-interactions",
         "member3": "Performance improvements, Database tuning, Caching",
         "milestone": ""},
        
        {"member1": "Final documentation, Video tutorials, User onboarding", 
         "member2": "Final documentation, Design system docs, Component docs",
         "member3": "Final documentation, API changelog, Runbook",
         "milestone": "Documentation Complete"},
        
        {"member1": "Handover preparation, Knowledge transfer, Training", 
         "member2": "Handover preparation, Design handoff, Asset delivery",
         "member3": "Handover preparation, Infrastructure docs, Access setup",
         "milestone": ""},
        
        {"member1": "Final review, Retrospective, Lessons learned", 
         "member2": "Final review, Retrospective, UI/UX insights",
         "member3": "Final review, Retrospective, Technical debt log",
         "milestone": ""},
        
        {"member1": "Project closure, Archive, Celebration! 🎉", 
         "member2": "Project closure, Portfolio update, Celebration! 🎉",
         "member3": "Project closure, Monitoring handoff, Celebration! 🎉",
         "milestone": "Project Complete! 🚀"},
    ]


def create_member1_tasks_sheet(wb):
    """Create Member 1 (Startup Portal) detailed tasks"""
    ws = wb.create_sheet("Member 1 - Startup")
    create_member_sheet(ws, "Member 1: Startup Portal Developer", get_member1_features())


def create_member2_tasks_sheet(wb):
    """Create Member 2 (Investor Frontend) detailed tasks"""
    ws = wb.create_sheet("Member 2 - Investor FE")
    create_member_sheet(ws, "Member 2: Investor Portal Frontend", get_member2_features())


def create_member3_tasks_sheet(wb):
    """Create Member 3 (Investor Backend) detailed tasks"""
    ws = wb.create_sheet("Member 3 - Investor BE")
    create_member_sheet(ws, "Member 3: Investor Portal Backend", get_member3_features())


def create_member_sheet(ws, title, features):
    """Create individual member task sheet"""
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    ws.merge_cells('A1:D1')
    
    headers = ["Feature", "Tasks", "Priority", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    
    row = 4
    for feature in features:
        ws.cell(row, 1, feature["name"])
        ws.cell(row, 2, feature["tasks"])
        ws.cell(row, 3, feature["priority"])
        ws.cell(row, 4, "Pending")
        
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        
        # Priority color coding
        priority_colors = {"High": "FF6B6B", "Medium": "FFA500", "Low": "4ECDC4"}
        ws.cell(row, 3).fill = PatternFill(start_color=priority_colors.get(feature["priority"], "FFFFFF"), 
                                           end_color=priority_colors.get(feature["priority"], "FFFFFF"), 
                                           fill_type="solid")
        
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15


def get_member1_features():
    """Startup Portal features"""
    return [
        {"name": "Authentication", "tasks": "Signup page, Login page, Password reset, Email verification", "priority": "High"},
        {"name": "Dashboard", "tasks": "Stats overview, Recent activity, Quick actions, Notifications", "priority": "High"},
        {"name": "Company Profile", "tasks": "Profile form, Edit mode, Logo upload, Social links", "priority": "High"},
        {"name": "Pitch Creation", "tasks": "Multi-step form, File upload, Preview, Draft saving", "priority": "High"},
        {"name": "Investor Discovery", "tasks": "Browse investors, Filter by criteria, Connection requests", "priority": "Medium"},
        {"name": "Messaging", "tasks": "Chat interface, File sharing, Conversation list, Notifications", "priority": "High"},
        {"name": "Analytics", "tasks": "Pitch views, Investor interest, Engagement metrics", "priority": "Medium"},
        {"name": "Settings", "tasks": "Profile settings, Password change, Notifications preferences", "priority": "Low"},
    ]


def get_member2_features():
    """Investor Frontend features"""
    return [
        {"name": "Dashboard", "tasks": "Portfolio stats, Recent pitches, Activity feed, Quick filters", "priority": "High"},
        {"name": "Browse Pitches", "tasks": "Grid/List view, Filters, Search, Sort options, Pitch cards", "priority": "High"},
        {"name": "Pitch Details", "tasks": "Modal view, Comments, Ratings, Data room access, Actions", "priority": "High"},
        {"name": "Portfolio", "tasks": "Investment list, Stats cards, Growth charts, Export", "priority": "High"},
        {"name": "Watchlist", "tasks": "Saved pitches, Add/remove, Notes, Reminders", "priority": "Medium"},
        {"name": "In Review", "tasks": "Pipeline view, Status updates, Notes, Decision tracking", "priority": "Medium"},
        {"name": "Messages", "tasks": "Chat UI, User search, File preview, Notifications", "priority": "High"},
        {"name": "Meetings", "tasks": "Calendar view, Schedule meeting, Time slots, Reminders", "priority": "Low"},
        {"name": "Reports", "tasks": "Export reports, PDF generation, Charts, Custom filters", "priority": "Low"},
    ]


def get_member3_features():
    """Investor Backend features"""
    return [
        {"name": "Authentication API", "tasks": "Register, Login, JWT, Password reset, Email service", "priority": "High"},
        {"name": "User Management", "tasks": "Profile CRUD, Role management, Permissions", "priority": "High"},
        {"name": "Pitch API", "tasks": "CRUD operations, Feed endpoint, Search, Filters, Sorting", "priority": "High"},
        {"name": "Investment API", "tasks": "Log investment, Portfolio stats, Analytics, Export", "priority": "High"},
        {"name": "Messaging API", "tasks": "Send/receive messages, File upload, WebSocket, Notifications", "priority": "High"},
        {"name": "Data Room", "tasks": "Document upload, Access control, Download tracking", "priority": "Medium"},
        {"name": "Connections", "tasks": "Request/accept/reject, Status management, Notifications", "priority": "Medium"},
        {"name": "Notifications", "tasks": "Create notifications, Mark read, Push notifications", "priority": "Medium"},
        {"name": "Infrastructure", "tasks": "Docker setup, Database migrations, Monitoring, Backups", "priority": "High"},
        {"name": "Security", "tasks": "Rate limiting, Input validation, SQL injection prevention", "priority": "High"},
    ]


def create_progress_tracker_sheet(wb):
    """Create progress tracking sheet"""
    ws = wb.create_sheet("Progress Tracker")
    
    ws['A1'] = "Weekly Progress Tracker"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    ws.merge_cells('A1:E1')
    
    headers = ["Week", "Planned Tasks", "Completed Tasks", "Completion %", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    
    weeks = [
        {"week": "Week 1", "planned": 21, "notes": "Foundation & Core Features"},
        {"week": "Week 2", "planned": 21, "notes": "Advanced Features"},
        {"week": "Week 3", "planned": 21, "notes": "Integration & Testing"},
        {"week": "Week 4", "planned": 27, "notes": "Deployment & Polish"},
    ]
    
    for idx, week in enumerate(weeks, 4):
        ws.cell(idx, 1, week["week"])
        ws.cell(idx, 2, week["planned"])
        ws.cell(idx, 3, 0)  # To be filled manually
        ws.cell(idx, 4, "=C{}/B{}*100".format(idx, idx))
        ws.cell(idx, 5, week["notes"])
        
        ws.cell(idx, 4).number_format = '0.00"%"'
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40


if __name__ == "__main__":
    print("🚀 Generating InvestorConnect Project Plan...")
    print("=" * 60)
    filename = create_project_plan()
    print("=" * 60)
    print(f"📊 Excel file created successfully!")
    print(f"📁 Location: {os.path.abspath(filename)}")
    print("\n✨ Project plan includes:")
    print("   • Project Overview")
    print("   • 30-Day Daily Task Breakdown")
    print("   • Member 1: Startup Portal Tasks")
    print("   • Member 2: Investor Frontend Tasks")
    print("   • Member 3: Investor Backend Tasks")
    print("   • Weekly Progress Tracker")
    print("\n💡 Open the file in Excel to view and track your project!")
